import openpyxl
import sys
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
import unicodedata
from difflib import SequenceMatcher

# Adicionar diretório src ao path para importar constantes
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from layout_generator import HEADER_FUZZY_THRESHOLD, MIN_FUZZY_THRESHOLD, DEFAULT_DATE
from config import EXPECTED_COLUMNS, HEADER_TOLERANCE


class ExcelReader:
    """Classe para ler e extrair tabelas de arquivos Excel."""
    
    EXPECTED_COLUMNS = EXPECTED_COLUMNS
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        
    def read_excel_file(self) -> None:
        """Carrega o arquivo Excel usando openpyxl com data_only=True para obter valores calculados."""
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        self.worksheet = self.workbook.active
        
    def normalize_header(self, value: Any) -> str:
        """
        Normaliza valor de célula para comparação de cabeçalhos.
        Remove espaços, quebras de linha, acentos e converte para maiúsculo.
        """
        if value is None:
            return ""
        
        # Converter para string e remover quebras de linha
        text = str(value).replace('\r', ' ').replace('\n', ' ')
        
        # Remover acentos
        text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
        
        # Converter para maiúsculo
        text = text.upper()
        
        # Remover espaços extras e normalizar
        text = ' '.join(text.split())
        
        return text.strip()
    
    def fuzzy_match(self, str1: str, str2: str, threshold: float = 0.7) -> bool:
        """
        Verifica se duas strings são similares usando fuzzy matching.
        Retorna True se a similaridade for maior que o threshold.
        """
        if not str1 or not str2:
            return False
        
        # Normalizar ambas as strings
        norm1 = self.normalize_header(str1)
        norm2 = self.normalize_header(str2)
        
        # Calcular similaridade usando SequenceMatcher
        similarity = SequenceMatcher(None, norm1, norm2).ratio()
        
        return similarity >= threshold
    
    def find_table_headers(self, start_row: int = 10) -> Optional[int]:
        """
        Localiza cabeçalhos de tabelas a partir de uma linha específica.
        Procura pelo cabeçalho específico com as colunas esperadas.
        """
        for row in range(start_row, self.worksheet.max_row + 1):
            headers_found = []
            for col in range(2, 19):  # Colunas B (2) a R (18)
                cell_value = self.worksheet.cell(row=row, column=col).value
                if cell_value:
                    headers_found.append(self.normalize_header(cell_value))
            
            # Verificar se encontrou pelo menos a maioria dos cabeçalhos usando fuzzy matching
            matches = 0
            for expected in self.EXPECTED_COLUMNS:
                # Verificar se algum cabeçalho encontrado faz fuzzy match com o esperado
                for found in headers_found:
                    if self.fuzzy_match(expected, found, threshold=HEADER_FUZZY_THRESHOLD):
                        matches += 1
                        break
            
            if matches >= len(self.EXPECTED_COLUMNS) - HEADER_TOLERANCE:  # Tolerância para colunas opcionais
                return row
                
        return None
    
    def extract_table_data(self, start_row: int) -> List[Dict[str, Any]]:
        """
        Extrai dados de uma tabela começando na linha do cabeçalho.
        Retorna quando a coluna B estiver vazia.
        Usa fuzzy matching para mapear colunas encontradas para nomes esperados.
        """
        # Primeiro, extrair os cabeçalhos normalizados
        headers_normalized = []
        for col in range(2, 19):  # Colunas B a R
            cell_value = self.worksheet.cell(row=start_row, column=col).value
            if cell_value:
                headers_normalized.append(self.normalize_header(cell_value))
            else:
                headers_normalized.append("")
        
        # Criar mapeamento usando fuzzy matching
        column_mapping = {}
        for idx, header_norm in enumerate(headers_normalized):
            if not header_norm:
                continue
            
            # Encontrar melhor correspondência com fuzzy matching
            best_match = None
            best_similarity = 0
            
            for expected in self.EXPECTED_COLUMNS:
                if self.fuzzy_match(header_norm, expected, threshold=MIN_FUZZY_THRESHOLD):
                    similarity = SequenceMatcher(None, header_norm, expected).ratio()
                    if similarity > best_similarity:
                        best_similarity = similarity
                        best_match = expected
            
            # Usar o nome esperado se houver match, senão usar o normalizado
            if best_match:
                column_mapping[idx] = best_match
            else:
                column_mapping[idx] = header_norm
        
        table_data = []
        
        # Extrair linhas de dados
        for row in range(start_row + 1, self.worksheet.max_row + 1):
            # Verificar se coluna B está vazia (fim da tabela)
            first_col_value = self.worksheet.cell(row=row, column=2).value
            if first_col_value is None or str(first_col_value).strip() == "":
                break
            
            # Extrair valores da linha usando o mapeamento de colunas
            row_data = {}
            for col_idx in range(2, 19):
                cell_value = self.worksheet.cell(row=row, column=col_idx).value
                # Usar o nome mapeado como chave
                mapped_name = column_mapping.get(col_idx - 2, "")
                if mapped_name:
                    row_data[mapped_name] = cell_value
            
            table_data.append(row_data)
        
        return table_data
    
    def detect_all_tables(self) -> List[Dict[str, Any]]:
        """
        Identifica todas as tabelas no arquivo.
        Busca cabeçalhos até 10 linhas após o fim da tabela anterior.
        """
        tables = []
        current_row = 10  # Começa a buscar a partir da linha 10
        
        while current_row <= self.worksheet.max_row:
            header_row = self.find_table_headers(current_row)
            
            if header_row is None:
                break
            
            # Extrair dados da tabela
            table_data = self.extract_table_data(header_row)
            
            if table_data:
                tables.append({
                    'header_row': header_row,
                    'data': table_data
                })
            
            # Buscar próxima tabela até 10 linhas após o fim desta
            current_row = header_row + len(table_data) + 1
            max_search_row = current_row + 10
            
            # Procurar próximo cabeçalho no intervalo
            next_header = None
            for search_row in range(current_row, min(max_search_row + 1, self.worksheet.max_row + 1)):
                if self.find_table_headers(search_row) == search_row:
                    next_header = search_row
                    break
            
            if next_header is None:
                break
            
            current_row = next_header
        
        return tables
    
    def convert_excel_date(self, serial_date) -> str:
        """
        Converte data serial do Excel para formato AAAAMMDD.
        """
        if serial_date is None:
            return DEFAULT_DATE
        
        try:
            if isinstance(serial_date, datetime):
                return serial_date.strftime("%Y%m%d")
            elif isinstance(serial_date, (int, float)):
                # Converter serial do Excel para datetime
                dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(serial_date) - 2)
                return dt.strftime("%Y%m%d")
            else:
                # Tentar converter string
                data_str = str(serial_date)
                if ' ' in data_str:
                    data_str = data_str.split(' ')[0]
                data_prestacao = data_str.replace('-', '').replace('/', '')
                if len(data_prestacao) == 8 and data_prestacao.isdigit():
                    return data_prestacao
                return DEFAULT_DATE
        except (ValueError, TypeError, OverflowError):
            return DEFAULT_DATE
    
    def close(self) -> None:
        """Fecha o arquivo Excel."""
        if self.workbook:
            self.workbook.close()
